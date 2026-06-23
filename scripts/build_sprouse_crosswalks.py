#!/usr/bin/env python3
"""Build structural Sprouse import summaries and crosswalks.

This script intentionally avoids method-performance summaries. It reads public
verification copies, records schemas and identifier coverage, and writes local
derived crosswalks under data/derived/.
"""

from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - environment guard
    raise SystemExit(
        "openpyxl is required to read the Sprouse materials workbooks. "
        "Install it or run in an environment that already has it."
    ) from exc


RESULT_HEADER_NAMES = {"participant", "subject"}
KEY_COLUMNS = (
    "participant",
    "subject",
    "survey",
    "item",
    "condition",
    "order",
    "label",
    "expected",
    "pattern",
)


@dataclass(frozen=True)
class CsvSpec:
    dataset: str
    role: str
    path: Path
    result_file: bool


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build structural Sprouse data summaries and crosswalks."
    )
    parser.add_argument(
        "--raw-root",
        type=Path,
        default=Path("/tmp/bard-data-check"),
        help="Directory containing Sprouse2013/ and Sprouse2017/ verification copies.",
    )
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=Path("data/derived/sprouse"),
        help="Output directory for ignored derived structural files.",
    )
    return parser.parse_args()


def decode_text(path: Path) -> tuple[str, str]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "mac_roman", "cp1252", "latin-1"):
        try:
            return raw.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace"), "utf-8-replace"


def first_csv_cell(line: str) -> str:
    try:
        return next(csv.reader([line]))[0].lstrip("\ufeff").strip()
    except (csv.Error, IndexError):
        return ""


def read_csv(path: Path, result_file: bool) -> tuple[list[dict[str, str]], list[str], int, str]:
    text, encoding = decode_text(path)
    lines = text.splitlines(True)
    start = 0
    if result_file:
        for index, line in enumerate(lines):
            if first_csv_cell(line) in RESULT_HEADER_NAMES:
                start = index
                break
        else:
            raise ValueError(f"Could not find data header in {path}")

    reader = csv.DictReader(lines[start:])
    return list(reader), list(reader.fieldnames or []), start, encoding


def write_csv(path: Path, fieldnames: Iterable[str], rows: Iterable[dict[str, object]]) -> None:
    field_list = list(fieldnames)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=field_list)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in field_list})


def normalize_item_to_condition(value: str) -> str:
    """Map item IDs to condition-like IDs without changing original IDs.

    Examples:
    - good.32.1.martin.20a.* -> 32.1.martin.20a.*
    - 32.1.martin.20a.*.01 -> 32.1.martin.20a.*
    """
    value = (value or "").strip()
    if value.startswith("good."):
        value = value[len("good.") :]
    parts = value.rsplit(".", 1)
    if len(parts) == 2 and parts[1].isdigit():
        return parts[0]
    return value


def compact_condition_key(value: str) -> str:
    """Secondary candidate key for documented spacing variants in IDs."""
    return normalize_item_to_condition(value).replace(" ", "")


def values(rows: list[dict[str, str]], key: str) -> set[str]:
    return {row[key] for row in rows if key in row and row.get(key)}


def row_by_key(rows: list[dict[str, str]], key: str) -> dict[str, dict[str, str]]:
    output: dict[str, dict[str, str]] = {}
    for row in rows:
        value = normalize_item_to_condition(row.get(key, ""))
        if value and value not in output:
            output[value] = row
    return output


def row_by_compact_key(rows: list[dict[str, str]], key: str) -> dict[str, dict[str, str]]:
    output: dict[str, dict[str, str]] = {}
    for row in rows:
        value = compact_condition_key(row.get(key, ""))
        if value and value not in output:
            output[value] = row
    return output


def schema_rows(specs: list[CsvSpec]) -> tuple[list[dict[str, object]], dict[str, list[dict[str, str]]]]:
    loaded: dict[str, list[dict[str, str]]] = {}
    summary: list[dict[str, object]] = []
    for spec in specs:
        rows, headers, skip_lines, encoding = read_csv(spec.path, spec.result_file)
        loaded[f"{spec.dataset}:{spec.role}"] = rows
        summary_row: dict[str, object] = {
            "dataset": spec.dataset,
            "role": spec.role,
            "source_path": str(spec.path),
            "encoding": encoding,
            "skip_lines": skip_lines,
            "rows": len(rows),
            "columns": "|".join(headers),
        }
        for key in KEY_COLUMNS:
            summary_row[f"unique_{key}"] = len(values(rows, key)) if key in headers else ""
        summary.append(summary_row)
    return summary, loaded


def load_workbook_rows(path: Path, sheet_name: str, header_row: int) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    headers = [
        "" if cell.value is None else str(cell.value).strip()
        for cell in next(sheet.iter_rows(min_row=header_row, max_row=header_row))
    ]
    output: list[dict[str, str]] = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        record = {
            headers[index]: "" if value is None else str(value).strip()
            for index, value in enumerate(row)
            if index < len(headers) and headers[index]
        }
        if any(record.values()):
            output.append(record)
    return output


def build_2013_crosswalks(out_dir: Path, loaded: dict[str, list[dict[str, str]]]) -> dict[str, int]:
    fc_conditions = loaded["2013:fc_conditions"]
    me_conditions = loaded["2013:me_conditions"]
    ls_conditions = loaded["2013:ls_conditions"]

    me_by_bad = row_by_key(me_conditions, "bad")
    ls_by_bad = row_by_key(ls_conditions, "bad")
    me_by_bad_compact = row_by_compact_key(me_conditions, "bad")
    ls_by_bad_compact = row_by_compact_key(ls_conditions, "bad")

    crosswalk_rows = []
    for index, fc_row in enumerate(fc_conditions, start=1):
        bad_base = normalize_item_to_condition(fc_row["bad"])
        bad_compact = compact_condition_key(fc_row["bad"])
        me_row_exact = me_by_bad.get(bad_base, {})
        ls_row_exact = ls_by_bad.get(bad_base, {})
        me_row_compact = me_by_bad_compact.get(bad_compact, {})
        ls_row_compact = ls_by_bad_compact.get(bad_compact, {})
        me_row = me_row_exact or me_row_compact
        ls_row = ls_row_exact or ls_row_compact
        match_basis = "unmatched"
        if me_row_exact and ls_row_exact:
            match_basis = "exact"
        elif me_row and ls_row:
            match_basis = "compact_spacing_variant"
        crosswalk_rows.append(
            {
                "pair_index": index,
                "fc_bad_condition": fc_row.get("bad", ""),
                "fc_good_condition": fc_row.get("good", ""),
                "fc_good_without_prefix": normalize_item_to_condition(fc_row.get("good", "")),
                "bad_condition_base": bad_base,
                "bad_condition_compact_key": bad_compact,
                "me_bad_item": me_row.get("bad", ""),
                "me_good_item": me_row.get("good", ""),
                "me_list": me_row.get("list", ""),
                "ls_bad_item": ls_row.get("bad", ""),
                "ls_good_item": ls_row.get("good", ""),
                "ls_list": ls_row.get("list", ""),
                "has_me_pair_exact": bool(me_row_exact),
                "has_ls_pair_exact": bool(ls_row_exact),
                "has_me_pair_compact": bool(me_row_compact),
                "has_ls_pair_compact": bool(ls_row_compact),
                "me_ls_pair_identical": bool(me_row and ls_row and me_row == ls_row),
                "match_basis": match_basis,
            }
        )

    write_csv(
        out_dir / "2013_condition_crosswalk.csv",
        (
            "pair_index",
            "fc_bad_condition",
            "fc_good_condition",
            "fc_good_without_prefix",
            "bad_condition_base",
            "bad_condition_compact_key",
            "me_bad_item",
            "me_good_item",
            "me_list",
            "ls_bad_item",
            "ls_good_item",
            "ls_list",
            "has_me_pair_exact",
            "has_ls_pair_exact",
            "has_me_pair_compact",
            "has_ls_pair_compact",
            "me_ls_pair_identical",
            "match_basis",
        ),
        crosswalk_rows,
    )

    method_condition_sets = {
        "fc_logistic_condition": values(loaded["2013:fc_logistic"], "condition"),
        "fc_signtest_condition": values(loaded["2013:fc_signtest"], "condition"),
        "me_condition": values(loaded["2013:me"], "condition"),
        "ls_condition": values(loaded["2013:ls"], "condition"),
        "fc_pair_bad_base": {normalize_item_to_condition(row["bad"]) for row in fc_conditions},
        "me_pair_bad_base": {normalize_item_to_condition(row["bad"]) for row in me_conditions},
        "ls_pair_bad_base": {normalize_item_to_condition(row["bad"]) for row in ls_conditions},
    }
    all_conditions = sorted(set().union(*method_condition_sets.values()))
    condition_presence = []
    for condition in all_conditions:
        row = {"condition_id": condition}
        for name, condition_set in method_condition_sets.items():
            row[f"in_{name}"] = condition in condition_set
        condition_presence.append(row)
    write_csv(
        out_dir / "2013_condition_presence.csv",
        ["condition_id"] + [f"in_{name}" for name in method_condition_sets],
        condition_presence,
    )

    return {
        "2013_condition_crosswalk_rows": len(crosswalk_rows),
        "2013_fc_pairs_with_me_pair_exact": sum(row["has_me_pair_exact"] for row in crosswalk_rows),
        "2013_fc_pairs_with_ls_pair_exact": sum(row["has_ls_pair_exact"] for row in crosswalk_rows),
        "2013_fc_pairs_with_me_pair_compact": sum(row["has_me_pair_compact"] for row in crosswalk_rows),
        "2013_fc_pairs_with_ls_pair_compact": sum(row["has_ls_pair_compact"] for row in crosswalk_rows),
        "2013_fc_pairs_still_unmatched": sum(row["match_basis"] == "unmatched" for row in crosswalk_rows),
    }


def build_2017_crosswalks(
    raw_root: Path, out_dir: Path, loaded: dict[str, list[dict[str, str]]]
) -> dict[str, int]:
    workbook_path = raw_root / "Sprouse2017" / "SA2017.materials.xlsx"
    label_rows = load_workbook_rows(workbook_path, "LI phenomena labels", 1)
    material_rows = load_workbook_rows(workbook_path, "materials", 1)

    label_crosswalk = [
        {
            "label": row.get("label", ""),
            "bad_condition": row.get("bad condition", ""),
            "good_condition": row.get("good condition", ""),
        }
        for row in label_rows
        if row.get("label")
    ]
    write_csv(
        out_dir / "2017_label_crosswalk.csv",
        ("label", "bad_condition", "good_condition"),
        label_crosswalk,
    )

    material_items = [{"item": row.get("item", "")} for row in material_rows if row.get("item")]
    write_csv(out_dir / "2017_material_items.csv", ("item",), material_items)

    method_items = {
        "me_item": values(loaded["2017:me"], "item"),
        "ls_item": values(loaded["2017:ls"], "item"),
        "fc_condition_as_item": values(loaded["2017:fc"], "condition"),
        "yn_item": values(loaded["2017:yn"], "item"),
        "materials_item": {row["item"] for row in material_items},
    }
    all_items = sorted(set().union(*method_items.values()))
    presence_rows = []
    for item in all_items:
        row = {"item": item}
        for name, item_set in method_items.items():
            row[f"in_{name}"] = item in item_set
        presence_rows.append(row)
    write_csv(
        out_dir / "2017_method_item_presence.csv",
        ["item"] + [f"in_{name}" for name in method_items],
        presence_rows,
    )

    me_ls_items = method_items["me_item"] | method_items["ls_item"]
    yn_only_rows = [
        {
            "item": item,
            "condition_values_in_yn": "|".join(
                sorted({row["condition"] for row in loaded["2017:yn"] if row.get("item") == item})
            ),
            "row_count": sum(1 for row in loaded["2017:yn"] if row.get("item") == item),
            "unique_subjects": len({row["subject"] for row in loaded["2017:yn"] if row.get("item") == item}),
            "order_values": "|".join(
                sorted({row["order"] for row in loaded["2017:yn"] if row.get("item") == item}, key=int)
            ),
            "in_materials": item in method_items["materials_item"],
        }
        for item in sorted(method_items["yn_item"] - me_ls_items)
    ]
    write_csv(
        out_dir / "2017_yn_only_items.csv",
        ("item", "condition_values_in_yn", "row_count", "unique_subjects", "order_values", "in_materials"),
        yn_only_rows,
    )

    yn_inclusion_rows = []
    for item in sorted(method_items["yn_item"]):
        exclude_as_non_material = item in {row["item"] for row in yn_only_rows}
        yn_inclusion_rows.append(
            {
                "item": item,
                "include_in_item_level_comparison": not exclude_as_non_material,
                "exclusion_reason": "non_material_control_or_practice_row"
                if exclude_as_non_material
                else "",
                "in_me": item in method_items["me_item"],
                "in_ls": item in method_items["ls_item"],
                "in_fc_condition": item in method_items["fc_condition_as_item"],
                "in_materials": item in method_items["materials_item"],
            }
        )
    write_csv(
        out_dir / "2017_yn_item_inclusion.csv",
        (
            "item",
            "include_in_item_level_comparison",
            "exclusion_reason",
            "in_me",
            "in_ls",
            "in_fc_condition",
            "in_materials",
        ),
        yn_inclusion_rows,
    )

    return {
        "2017_label_crosswalk_rows": len(label_crosswalk),
        "2017_material_item_rows": len(material_items),
        "2017_yn_only_items": len(yn_only_rows),
        "2017_yn_items_included_for_item_level_comparison": sum(
            row["include_in_item_level_comparison"] for row in yn_inclusion_rows
        ),
    }


def write_analysis_rules(out_dir: Path) -> None:
    rows = [
        {
            "dataset": "2013",
            "rule": "forced_choice_primary_representation",
            "decision": "FC.signtest.csv",
            "status": "primary",
            "rationale": (
                "Forced choice is a pairwise task; the sign-test file is already "
                "at the pair/expected-pattern level and aligns with the pair-level "
                "ME/LS condition crosswalk."
            ),
        },
        {
            "dataset": "2013",
            "rule": "forced_choice_sensitivity_representation",
            "decision": "FC.logistic.csv",
            "status": "sensitivity",
            "rationale": (
                "The logistic file preserves individual item-level binary rows; "
                "use only as a named sensitivity representation because it changes "
                "the forced-choice unit from pair-level decisions to item-level rows."
            ),
        },
        {
            "dataset": "2017",
            "rule": "yes_no_item_level_exclusions",
            "decision": "exclude item IDs B, G, and M from item-level comparisons",
            "status": "primary",
            "rationale": (
                "These IDs are absent from the materials workbook, appear at repeated "
                "early order positions for every subject, and do not align with the "
                "ME/LS/FC item universe."
            ),
        },
    ]
    write_csv(
        out_dir / "analysis_rules.csv",
        ("dataset", "rule", "decision", "status", "rationale"),
        rows,
    )


def write_readme(out_dir: Path) -> None:
    text = """# Sprouse Derived Structural Files

Generated by `scripts/build_sprouse_crosswalks.py`.

These files are local structural derivatives only. They are ignored by Git and
should not be treated as publication-ready novel analysis outputs while the
Sprouse reuse question is pending.

The script does not compute method rankings, convergence rates, means, model
fits, or any other substantive outcome summaries.

`analysis_rules.csv` records pre-analysis representation and inclusion decisions
that future scripts should read before computing any outcomes.
"""
    (out_dir / "README.md").write_text(text, encoding="utf-8")


def main() -> int:
    args = parse_args()
    raw_root = args.raw_root
    out_dir = args.out_dir

    specs = [
        CsvSpec("2013", "fc_logistic", raw_root / "Sprouse2013/SSA.data/FC experiment/FC.logistic.csv", True),
        CsvSpec("2013", "fc_signtest", raw_root / "Sprouse2013/SSA.data/FC experiment/FC.signtest.csv", True),
        CsvSpec("2013", "fc_conditions", raw_root / "Sprouse2013/SSA.data/FC experiment/conditions.csv", False),
        CsvSpec("2013", "me", raw_root / "Sprouse2013/SSA.data/ME experiment/LI.me.results.csv", True),
        CsvSpec("2013", "me_conditions", raw_root / "Sprouse2013/SSA.data/ME experiment/LI.conditions.csv", False),
        CsvSpec("2013", "ls", raw_root / "Sprouse2013/SSA.data/LS experiment/LI.ls.results.csv", True),
        CsvSpec("2013", "ls_conditions", raw_root / "Sprouse2013/SSA.data/LS experiment/LI.conditions.csv", False),
        CsvSpec("2017", "me", raw_root / "Sprouse2017/SA2017.data/ME.results.csv", True),
        CsvSpec("2017", "ls", raw_root / "Sprouse2017/SA2017.data/LS.results.csv", True),
        CsvSpec("2017", "fc", raw_root / "Sprouse2017/SA2017.data/FC.results.csv", True),
        CsvSpec("2017", "yn", raw_root / "Sprouse2017/SA2017.data/YN.results.csv", True),
    ]

    missing = [spec.path for spec in specs if not spec.path.exists()]
    missing.append(raw_root / "Sprouse2017/SA2017.materials.xlsx")
    missing = [path for path in missing if not path.exists()]
    if missing:
        for path in missing:
            print(f"Missing input: {path}", file=sys.stderr)
        return 2

    out_dir.mkdir(parents=True, exist_ok=True)
    summary, loaded = schema_rows(specs)
    write_csv(
        out_dir / "schema_summary.csv",
        (
            "dataset",
            "role",
            "source_path",
            "encoding",
            "skip_lines",
            "rows",
            "columns",
            *[f"unique_{key}" for key in KEY_COLUMNS],
        ),
        summary,
    )

    report = {}
    report.update(build_2013_crosswalks(out_dir, loaded))
    report.update(build_2017_crosswalks(raw_root, out_dir, loaded))
    write_analysis_rules(out_dir)
    write_readme(out_dir)

    print(f"Wrote structural Sprouse derivatives to {out_dir}")
    for key, value in report.items():
        print(f"{key}: {value}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
